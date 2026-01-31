"""
CMHC Rental Market Survey (RMR) Victoria 2025 — load average rents by zone from Excel.
Source: rmr-victoria-2025-en.xlsx, Table 1.1.2 (Private Apartment Average Rents by Zone and Bedroom Type).
"""
from pathlib import Path
from typing import Any, Dict, List, Optional

# Project root (core/rmr_data.py -> parent -> parent)
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_RMR_EXCEL = _PROJECT_ROOT / "rmr-victoria-2025-en.xlsx"

# Table 1.1.2 layout: row 6 = header, row 7+ = data. Col 0 = Zone name.
# Oct-25 values: Studio col 3, 1BR col 7, 2BR col 11, 3BR+ col 15, Total col 19
_COL_OCT25 = {"studio": 3, "1bed": 7, "2bed": 11, "3bed": 15, "total": 19}
_BEDROOM_KEYS = ["studio", "1bed", "2bed", "3bed", "total"]

_ZONES_CACHE: Optional[List[Dict[str, Any]]] = None
_CMA_2BED: Optional[float] = None
_CITY_OF_VICTORIA_2BED: Optional[float] = None


def _parse_cell(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val) if val != "" else None
    s = str(val).strip().replace(",", "")
    if s in ("", "**", "–", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _load_rmr_table() -> None:
    global _ZONES_CACHE, _CMA_2BED, _CITY_OF_VICTORIA_2BED
    if _ZONES_CACHE is not None:
        return
    _ZONES_CACHE = []
    _CMA_2BED = None
    _CITY_OF_VICTORIA_2BED = None
    if not _RMR_EXCEL.exists():
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_RMR_EXCEL, read_only=True, data_only=True)
        sheet = None
        for sn in wb.sheetnames:
            if "1.1.2" in sn:
                sheet = wb[sn]
                break
        if sheet is None:
            wb.close()
            return
        rows = list(sheet.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 8:
            return
        for row in rows[7:20]:
            if not row or row[0] is None:
                continue
            zone_name = str(row[0]).strip()
            if not zone_name or zone_name.startswith("§") or zone_name.startswith("Quality") or zone_name.startswith("Source") or zone_name.startswith("©"):
                continue
            rents = {}
            for key, col in _COL_OCT25.items():
                rents[key] = _parse_cell(row[col]) if col < len(row) else None
            _ZONES_CACHE.append({
                "zone": zone_name,
                "studio": rents.get("studio"),
                "1bed": rents.get("1bed"),
                "2bed": rents.get("2bed"),
                "3bed": rents.get("3bed"),
                "total": rents.get("total"),
            })
            if "Victoria CMA" in zone_name:
                _CMA_2BED = rents.get("2bed")
            if "City of Victoria (Zones 1-4)" in zone_name:
                _CITY_OF_VICTORIA_2BED = rents.get("2bed")
    except Exception:
        pass


def get_rmr_zones() -> List[Dict[str, Any]]:
    """Return list of zone dicts: zone, studio, 1bed, 2bed, 3bed, total (Oct-25 avg rents)."""
    _load_rmr_table()
    return list(_ZONES_CACHE) if _ZONES_CACHE else []


def get_cma_2bed_rent() -> Optional[float]:
    """Victoria CMA 2-bedroom average rent (Oct 2025)."""
    _load_rmr_table()
    return _CMA_2BED


def get_city_of_victoria_2bed_rent() -> Optional[float]:
    """City of Victoria (Zones 1-4) 2-bedroom average rent (Oct 2025)."""
    _load_rmr_table()
    return _CITY_OF_VICTORIA_2BED


def get_rent_by_zone(zone_name: str, bedroom: str = "2bed") -> Optional[float]:
    """Return average rent for a zone and bedroom type (studio, 1bed, 2bed, 3bed, total)."""
    _load_rmr_table()
    if not _ZONES_CACHE or bedroom not in _BEDROOM_KEYS:
        return None
    for z in _ZONES_CACHE:
        if z["zone"] == zone_name:
            return z.get(bedroom)
    return None
